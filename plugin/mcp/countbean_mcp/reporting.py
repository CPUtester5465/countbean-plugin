"""Report generation for a Countbean book: styled HTML and Excel (.xlsx).

Both formats are built from the same BQL queries so the numbers always agree.
"""
from __future__ import annotations

import datetime
import html
import re
from pathlib import Path

from .ledger import Ledger, LedgerError

_NUM = re.compile(r"-?\d[\d,]*\.?\d*")


def _amount(cell: str) -> float:
    m = _NUM.search(cell or "")
    return float(m.group(0).replace(",", "")) if m else 0.0


def _operating_currency(ledger: Ledger) -> str:
    for line in ledger.main.read_text().splitlines():
        if 'operating_currency' in line:
            parts = line.split('"')
            if len(parts) >= 4:
                return parts[3]
    return "USD"


def _section(ledger: Ledger, root: str, ccy: str, date: str) -> list[dict]:
    # `date` is a bare BQL date literal (e.g. 2026-08-02); filter to on-or-before it.
    clause = f" AND date <= {date}" if date else ""
    bql = (
        f"SELECT account, sum(convert(position, '{ccy}')) AS balance "
        f"WHERE account ~ '^{root}'{clause} GROUP BY account ORDER BY account"
    )
    try:
        rows = ledger.query(bql)
    except LedgerError:
        rows = []
    out = []
    for r in rows:
        bal_key = next((k for k in r if k.lower().startswith("bal")), None)
        bal = r.get(bal_key, "") if bal_key else ""
        out.append({"account": r.get("account", ""), "amount": _amount(bal)})
    return out


def collect(ledger: Ledger, as_of: str | None = None) -> dict:
    ccy = _operating_currency(ledger)
    date = as_of or datetime.date.today().isoformat()
    assets = _section(ledger, "Assets", ccy, date)
    liabilities = _section(ledger, "Liabilities", ccy, date)
    equity = _section(ledger, "Equity", ccy, date)
    income = _section(ledger, "Income", ccy, date)
    expenses = _section(ledger, "Expenses", ccy, date)
    txn_clause = f"WHERE date <= {date} " if date else ""
    try:
        txns = ledger.query(
            f"SELECT date, payee, narration, account, convert(position, '{ccy}') AS amount "
            f"{txn_clause}ORDER BY date DESC, narration LIMIT 100"
        )
    except LedgerError:
        txns = []
    total = lambda s: round(sum(x["amount"] for x in s), 2)
    return {
        "title": _title(ledger),
        "currency": ccy,
        "as_of": date,
        "generated": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        "assets": assets, "liabilities": liabilities, "equity": equity,
        "income": income, "expenses": expenses, "transactions": txns,
        "totals": {
            "assets": total(assets),
            "liabilities": total(liabilities),
            "equity": total(equity),
            "income": total(income),
            "expenses": total(expenses),
            "net_worth": round(total(assets) + total(liabilities), 2),
            "net_income": round(-(total(income) + total(expenses)), 2),
        },
    }


def _title(ledger: Ledger) -> str:
    for line in ledger.main.read_text().splitlines():
        if line.startswith('option "title"'):
            parts = line.split('"')
            if len(parts) >= 4:
                return parts[3]
    return "Countbean"


# ---------------------------------------------------------------- HTML -------
def generate_html(ledger: Ledger, out_path: str, as_of: str | None = None) -> str:
    d = collect(ledger, as_of)
    c = d["currency"]

    def money(n: float) -> str:
        sign = "neg" if n < 0 else "pos" if n > 0 else "zero"
        return f'<span class="m {sign}">{n:,.2f}</span>'

    def rows(section: list[dict]) -> str:
        if not section:
            return '<tr><td class="acct empty" colspan="2">No entries yet</td></tr>'
        return "".join(
            f'<tr><td class="acct">{html.escape(x["account"])}</td>'
            f'<td class="num">{money(x["amount"])}</td></tr>'
            for x in section
        )

    def txn_rows() -> str:
        if not d["transactions"]:
            return '<tr><td class="empty" colspan="4">No transactions yet</td></tr>'
        out = []
        for t in d["transactions"]:
            amt_key = next((k for k in t if k.lower().startswith("amount")), None)
            out.append(
                "<tr>"
                f'<td class="date">{html.escape(t.get("date",""))}</td>'
                f'<td>{html.escape(t.get("payee","") or "")}</td>'
                f'<td>{html.escape(t.get("narration","") or "")}</td>'
                f'<td class="acct">{html.escape(t.get("account",""))}</td>'
                f'<td class="num">{money(_amount(t.get(amt_key,"")))}</td>'
                "</tr>"
            )
        return "".join(out)

    t = d["totals"]
    doc = _HTML_TEMPLATE.format(
        title=html.escape(d["title"]),
        as_of=d["as_of"],
        generated=d["generated"],
        ccy=c,
        net_worth=money(t["net_worth"]),
        net_income=money(t["net_income"]),
        assets_total=money(t["assets"]),
        liabilities_total=money(t["liabilities"]),
        assets_rows=rows(d["assets"]),
        liabilities_rows=rows(d["liabilities"]),
        equity_rows=rows(d["equity"]),
        income_rows=rows(d["income"]),
        expenses_rows=rows(d["expenses"]),
        income_total=money(t["income"]),
        expenses_total=money(t["expenses"]),
        txn_rows=txn_rows(),
    )
    Path(out_path).write_text(doc)
    return out_path


# ---------------------------------------------------------------- Excel ------
def generate_xlsx(ledger: Ledger, out_path: str, as_of: str | None = None) -> str:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError as e:  # pragma: no cover
        raise LedgerError("openpyxl not installed — `pip install openpyxl`.") from e

    d = collect(ledger, as_of)
    wb = Workbook()
    head_fill = PatternFill("solid", fgColor="146B4A")
    head_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14, color="146B4A")
    money_fmt = '#,##0.00;[Red]-#,##0.00'

    def sheet(name, sections):
        ws = wb.create_sheet(name)
        ws["A1"] = f'{d["title"]} — {name}'
        ws["A1"].font = title_font
        ws["A2"] = f'As of {d["as_of"]}  ·  {d["currency"]}'
        ws["A2"].font = Font(italic=True, color="6D7D74")
        r = 4
        for label, rows_ in sections:
            ws.cell(r, 1, label).font = Font(bold=True)
            r += 1
            ws.cell(r, 1, "Account").font = head_font
            ws.cell(r, 2, "Balance").font = head_font
            for col in (1, 2):
                ws.cell(r, col).fill = head_fill
            r += 1
            subtotal = 0.0
            for x in rows_:
                ws.cell(r, 1, x["account"])
                cell = ws.cell(r, 2, x["amount"])
                cell.number_format = money_fmt
                cell.alignment = Alignment(horizontal="right")
                subtotal += x["amount"]
                r += 1
            ws.cell(r, 1, f"Total {label}").font = Font(bold=True)
            tc = ws.cell(r, 2, round(subtotal, 2))
            tc.font = Font(bold=True)
            tc.number_format = money_fmt
            r += 2
        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 18
        return ws

    sheet("Balance Sheet",
          [("Assets", d["assets"]), ("Liabilities", d["liabilities"]), ("Equity", d["equity"])])
    sheet("Income Statement",
          [("Income", d["income"]), ("Expenses", d["expenses"])])

    ws = wb.create_sheet("Transactions")
    headers = ["Date", "Payee", "Narration", "Account", "Amount"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = head_font
        cell.fill = head_fill
    for i, tr in enumerate(d["transactions"], start=2):
        amt_key = next((k for k in tr if k.lower().startswith("amount")), None)
        ws.cell(i, 1, tr.get("date", ""))
        ws.cell(i, 2, tr.get("payee", "") or "")
        ws.cell(i, 3, tr.get("narration", "") or "")
        ws.cell(i, 4, tr.get("account", ""))
        c = ws.cell(i, 5, _amount(tr.get(amt_key, "")))
        c.number_format = money_fmt
    widths = [12, 22, 30, 34, 14]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = w

    wb.remove(wb["Sheet"])  # drop default
    wb.save(out_path)
    return out_path


_HTML_TEMPLATE = """<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{title} — Report</title>
<style>
  :root {{
    --paper:#eef1ea; --card:#f7f9f3; --ink:#16211a; --soft:#40524a; --faint:#6d7d74;
    --line:#cdd6c8; --pine:#146b4a; --pos:#147a52; --neg:#b04a2f;
    --mono:ui-monospace,"SF Mono",Menlo,Consolas,monospace;
    --sans:ui-sans-serif,system-ui,-apple-system,"Segoe UI",Roboto,sans-serif;
  }}
  @media (prefers-color-scheme:dark) {{
    :root {{ --paper:#0e1512; --card:#14201a; --ink:#e7ede4; --soft:#b3c1b6; --faint:#7d8f83;
      --line:#26332b; --pine:#2fb37c; --pos:#43c088; --neg:#e08363; }}
  }}
  * {{ box-sizing:border-box; }}
  body {{ margin:0; font-family:var(--sans); background:var(--paper); color:var(--ink);
    line-height:1.55; padding:clamp(20px,5vw,56px); }}
  .wrap {{ max-width:900px; margin:0 auto; }}
  header {{ display:flex; justify-content:space-between; align-items:flex-end; flex-wrap:wrap;
    gap:12px; border-bottom:2px solid var(--pine); padding-bottom:16px; margin-bottom:8px; }}
  h1 {{ font-family:var(--mono); font-size:clamp(22px,3vw,30px); margin:0; letter-spacing:-.02em; }}
  .meta {{ font-family:var(--mono); font-size:12.5px; color:var(--faint); text-align:right; }}
  .kpis {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(160px,1fr)); gap:14px;
    margin:26px 0; }}
  .kpi {{ background:var(--card); border:1px solid var(--line); border-radius:8px; padding:16px 18px; }}
  .kpi .l {{ font-family:var(--mono); font-size:11px; letter-spacing:.1em; text-transform:uppercase;
    color:var(--faint); }}
  .kpi .v {{ font-family:var(--mono); font-size:26px; font-variant-numeric:tabular-nums; margin-top:6px; }}
  h2 {{ font-family:var(--mono); font-size:15px; letter-spacing:.05em; text-transform:uppercase;
    color:var(--pine); margin:34px 0 12px; }}
  .cols {{ display:grid; grid-template-columns:1fr 1fr; gap:22px; }}
  @media (max-width:640px) {{ .cols {{ grid-template-columns:1fr; }} }}
  table {{ width:100%; border-collapse:collapse; background:var(--card);
    border:1px solid var(--line); border-radius:8px; overflow:hidden; }}
  caption {{ text-align:left; font-family:var(--mono); font-size:12px; color:var(--faint);
    padding:10px 12px; background:transparent; }}
  td, th {{ padding:8px 12px; border-bottom:1px solid var(--line); text-align:left; font-size:14px; }}
  tr:last-child td {{ border-bottom:none; }}
  .acct {{ font-family:var(--mono); font-size:13px; }}
  .num, .m {{ font-family:var(--mono); text-align:right; font-variant-numeric:tabular-nums; }}
  td.num {{ text-align:right; }}
  .m.pos {{ color:var(--pos); }} .m.neg {{ color:var(--neg); }} .m.zero {{ color:var(--faint); }}
  .total td {{ font-weight:700; border-top:2px solid var(--line); }}
  .empty {{ color:var(--faint); font-style:italic; }}
  .scroll {{ overflow-x:auto; }}
  footer {{ margin-top:40px; font-family:var(--mono); font-size:11.5px; color:var(--faint);
    text-align:center; }}
</style></head>
<body><div class="wrap">
  <header>
    <h1>{title}</h1>
    <div class="meta">Report as of {as_of}<br>Generated {generated} · {ccy}</div>
  </header>

  <div class="kpis">
    <div class="kpi"><div class="l">Net worth</div><div class="v">{net_worth}</div></div>
    <div class="kpi"><div class="l">Assets</div><div class="v">{assets_total}</div></div>
    <div class="kpi"><div class="l">Liabilities</div><div class="v">{liabilities_total}</div></div>
    <div class="kpi"><div class="l">Net income</div><div class="v">{net_income}</div></div>
  </div>

  <h2>Balance Sheet</h2>
  <div class="cols">
    <table><caption>Assets</caption>{assets_rows}</table>
    <table><caption>Liabilities &amp; Equity</caption>{liabilities_rows}{equity_rows}</table>
  </div>

  <h2>Income Statement</h2>
  <div class="cols">
    <table><caption>Income</caption>{income_rows}
      <tr class="total"><td class="acct">Total income</td><td class="num">{income_total}</td></tr></table>
    <table><caption>Expenses</caption>{expenses_rows}
      <tr class="total"><td class="acct">Total expenses</td><td class="num">{expenses_total}</td></tr></table>
  </div>

  <h2>Recent Transactions</h2>
  <div class="scroll"><table>
    <tr><th>Date</th><th>Payee</th><th>Narration</th><th>Account</th><th class="num">Amount</th></tr>
    {txn_rows}
  </table></div>

  <footer>Generated by Countbean · plain-text accounting, maintained by AI</footer>
</div></body></html>
"""
